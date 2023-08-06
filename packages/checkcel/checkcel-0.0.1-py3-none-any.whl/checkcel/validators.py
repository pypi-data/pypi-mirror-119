from email_validator import validate_email, EmailNotValidError
import requests

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, column_index_from_string
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from urllib.parse import quote_plus
from dateutil import parser

from collections import defaultdict

from checkcel.exceptions import ValidationException, BadValidatorException


class Validator(object):
    """ Generic Validator class """

    def __init__(self, empty_ok=None):
        self.invalid_dict = defaultdict(set)
        self.fail_count = 0
        self.empty_ok = empty_ok

    @property
    def bad(self):
        raise NotImplementedError

    def validate(self, field, row_number, row):
        """ Validate the given field. Also is given the row context """
        raise NotImplementedError

    def generate(self, column):
        """ Generate an openpyxl Datavalidation entity. Pass the column for custom formulas"""
        raise NotImplementedError

    def describe(self, column_name):
        """ Return a line of text describing allowed values"""
        raise NotImplementedError

    def _set_empty_ok(self, empty_ok_template):
        # Override with template value if it was not set (default to None)
        if self.empty_ok is None:
            self.empty_ok = empty_ok_template


class NoValidator(Validator):
    """ No check"""

    def __init__(self, **kwargs):
        super(NoValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        pass

    def generate(self, column):
        return None

    def describe(self, column_name):
        return "{} : Free text".format(column_name)

    @property
    def bad(self):
        return self.invalid_dict


class TextValidator(Validator):
    """ Default validator : will only check if not empty"""

    def __init__(self, **kwargs):
        super(TextValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        if not field and not self.empty_ok:
            raise ValidationException(
                "Field cannot be empty"
            )

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        return None

    def describe(self, column_name):
        return "{} : Free text".format(column_name)


class CastValidator(Validator):
    """ Validates that a field can be cast to a float """

    def __init__(self, min=None, max=None, **kwargs):
        super(CastValidator, self).__init__(**kwargs)
        self.min = min
        self.max = max

    def validate(self, field, row_number, row={}):
        try:
            if field or not self.empty_ok:
                field = self.cast(field)
                if self.min is not None and field < self.min:
                    self.invalid_dict["invalid_set"].add(field)
                    self.invalid_dict["invalid_rows"].add(row_number)
                    raise ValidationException("{} is below min value {}".format(field, self.min))
                if self.max is not None and field > self.max:
                    self.invalid_dict["invalid_set"].add(field)
                    self.invalid_dict["invalid_rows"].add(row_number)
                    raise ValidationException("{} is above max value {}".format(field, self.max))

        except ValueError as e:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        params = {"type": self.type}
        if (self.min is not None and self.max is not None):
            params["formula1"] = self.min
            params["formula2"] = self.max
            params["operator"] = "between"
        elif self.min is not None:
            params["formula1"] = self.min
            params["operator"] = "greaterThanOrEqual"
        elif self.max is not None:
            params["formula1"] = self.max
            params["operator"] = "lessThanOrEqual"
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : {} number".format(column_name, self.type.capitalize())
        if (self.min is not None and self.max is not None):
            text += " ({} - {})".format(self.min, self.max)
        elif self.min is not None:
            text += " >= {}".format(self.min)
        elif self.max is not None:
            text += " <= {}".format(self.max)
        return text


class FloatValidator(CastValidator):
    """ Validates that a field can be cast to a float """

    def __init__(self, **kwargs):
        super(FloatValidator, self).__init__(**kwargs)
        self.cast = float
        self.type = "decimal"


class IntValidator(CastValidator):
    """ Validates that a field can be cast to an int """

    def __init__(self, **kwargs):
        super(IntValidator, self).__init__(**kwargs)
        self.cast = int
        self.type = "whole"


class SetValidator(Validator):
    """ Validates that a field is in the given set of values """

    def __init__(self, valid_values=set(), **kwargs):
        super(SetValidator, self).__init__(**kwargs)
        self.valid_values = set(valid_values)
        if self.empty_ok:
            self.valid_values.add("")

    def validate(self, field, row_number, row={}):
        if field not in self.valid_values:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(
                "'{}' is invalid".format(field)
            )

    def _set_empty_ok(self, empty_ok_template):
        # Override with template value if it was not set (default to None)
        if self.empty_ok is None:
            self.empty_ok = empty_ok_template
        if self.empty_ok:
            self.valid_values.add("")

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, column_name="", additional_column=None, additional_worksheet=None):
        # If total length > 256 : need to use cells on another sheet
        if additional_column and additional_worksheet:
            params = {"type": "list"}
            cell = additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=column_name)
            cell.font = Font(color="FF0000", bold=True)
            row = 2
            for term in self.valid_values:
                additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=term)
                row += 1
            params["formula1"] = "{}!${}$2:${}${}".format(quote_sheetname(additional_worksheet.title), additional_column, additional_column, row - 1)
        else:
            params = {"type": "list"}
            values = ",".join(self.valid_values)
            params["formula1"] = '"{}"'.format(values)
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : ({})".format(column_name, ", ".join(self.valid_values))


class LinkedSetValidator(Validator):
    """ Validates that a field is in the given set of values """

    def __init__(self, linked_column, valid_values, **kwargs):
        super(LinkedSetValidator, self).__init__(**kwargs)
        self.valid_values = valid_values
        self.linked_column = linked_column
        self.column_check = False

    def _precheck_unique_with(self, row):
        if self.linked_column not in row.keys():
            raise BadValidatorException("Linked column {} is not in file columns".format(self.linked_column))
        self.column_check = True

    def validate(self, field, row_number, row):
        if not self.column_check:
            self._precheck_unique_with(row)
        if field == "" and self.empty_ok:
            return
        related_column_value = row[self.linked_column]
        if not related_column_value:
            self.invalid_dict["invalid_rows"].add(row_number)
            self.invalid_dict["invalid_set"].add("Invalid linked column value: ''")
            raise ValidationException("Linked column {} is empty".format(self.linked_column))
        if related_column_value not in self.valid_values.keys():
            self.invalid_dict["invalid_set"].add("Invalid linked column value: {}".format(related_column_value))
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("Linked column {} value {} is not in valid values".format(self.linked_column, related_column_value))
        if field not in self.valid_values[related_column_value]:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("Value {} is not in allowed values".format(field))

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, set_columns, column_name, additional_column, additional_worksheet, workbook):
        if self.linked_column not in set_columns:
            # TODO raise warning
            return None
        params = {"type": "list"}
        additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=column_name).font = Font(color="FF0000", bold=True)
        row = 2
        row_dict = {}
        for key, value in self.valid_values.items():
            additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=key).font = Font(color="FF0000", italic=True)
            row += 1
            row_dict[key] = {'min': row}
            for val in value:
                additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=val)
                row += 1
            row_dict[key]['max'] = row - 1
        for key, values in row_dict.items():
            new_range = DefinedName(key, attr_text='{}!${}${}:${}${}'.format(quote_sheetname(additional_worksheet.title), additional_column, values['min'], additional_column, values['max']))
            workbook.defined_names.append(new_range)
        params["formula1"] = "=INDIRECT(${}2)".format(set_columns[self.linked_column])
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Linked values to column {}".format(column_name, self.linked_column)


class DateValidator(Validator):
    """ Validates that a field is a Date """

    def __init__(self, day_first=True, **kwargs):
        super(DateValidator, self).__init__(**kwargs)
        self.day_first = day_first

    def validate(self, field, row_number, row={}):
        try:
            if field or not self.empty_ok:
                # Pandas auto convert fields into dates (ignoring the parse_dates=False)
                field = str(field)
                parser.parse(field, dayfirst=self.day_first).date()

        except parser.ParserError as e:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column=None, additional_worksheet=None):
        # GreaterThanOrEqual for validity with ODS.
        dv = DataValidation(type="date", formula1='12/30/1899', operator='greaterThanOrEqual')
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Date".format(column_name)


class EmailValidator(Validator):
    """ Validates that a field is in the given set """

    def __init__(self, **kwargs):
        super(EmailValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        if field or not self.empty_ok:
            try:
                validate_email(field)
            except EmailNotValidError as e:
                self.invalid_dict["invalid_set"].add(field)
                self.invalid_dict["invalid_rows"].add(row_number)
                raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, ontology_column=None):
        params = {"type": "custom"}
        params["formula1"] = '=ISNUMBER(MATCH("*@*.?*",{}2,0))'.format(column)
        dv = DataValidation(**params)
        dv.error = 'Value must be an email'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Email".format(column_name)


class OntologyValidator(Validator):
    """ Validates that a field is in the given set """

    def __init__(self, ontology, root_term="", **kwargs):
        super(OntologyValidator, self).__init__(**kwargs)
        self.validated_terms = set()
        self.ontology = ontology
        self.root_term = root_term
        self.root_term_iri = ""

        is_ontology, self.root_term_iri = _validate_ontology(ontology, self.root_term)
        if not is_ontology:
            raise BadValidatorException("'{}' is not a valid ontology".format(self.ontology))
        if self.root_term and not self.root_term_iri:
            raise BadValidatorException("'{}' is not a valid root term for ontology {}".format(self.root_term, self.ontology))

    def validate(self, field, row_number, row={}):
        if field == "" and self.empty_ok:
            return

        if field in self.invalid_dict["invalid_set"]:
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("{} is not an ontological term".format(field))

        if field not in self.validated_terms:
            ontological_term = _validate_ontological_term(field, self.ontology, self.root_term_iri)
            if not ontological_term:
                self.invalid_dict["invalid_set"].add(field)
                self.invalid_dict["invalid_rows"].add(row_number)
                raise ValidationException("{} is not an ontological term".format(field))
            self.validated_terms.add(field)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column, additional_worksheet):
        terms = _get_ontological_terms(self.ontology, root_term_iri=self.root_term_iri)
        cell = additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=self.ontology)
        cell.font = Font(color="FF0000", bold=True)
        row = 2
        for term in terms:
            additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=term)
            row += 1

        params = {"type": "list"}
        params["formula1"] = "{}!${}$2:${}${}".format(quote_sheetname(additional_worksheet.title), additional_column, additional_column, row - 1)
        dv = DataValidation(**params)
        dv.error = 'Value must be an ontological term'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Ontological term from {} ontology.".format(column_name, self.ontology)
        if self.root_term:
            text += " Root term is : {}".format(self.root_term)
        return text


class UniqueValidator(Validator):
    """ Validates that a field is unique within the file """

    def __init__(self, unique_with=[], **kwargs):
        super(UniqueValidator, self).__init__(**kwargs)
        self.unique_values = set()
        self.unique_with = unique_with
        self.unique_check = False

    def _precheck_unique_with(self, row):
        extra = set(self.unique_with) - set(row.keys())
        if extra:
            raise BadValidatorException(extra)
        self.unique_check = True

    def validate(self, field, row_number, row={}):
        if field == "" and self.empty_ok:
            return
        if self.unique_with and not self.unique_check:
            self._precheck_unique_with(row)

        key = tuple([field] + [row[k] for k in self.unique_with])
        if key not in self.unique_values:
            self.unique_values.add(key)
        else:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            if self.unique_with:
                raise ValidationException(
                    "'{}' is already in the column (unique with: {})".format(
                        field, key[1:]
                    )
                )
            else:
                raise ValidationException("'{}' is already in the column".format(field))

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, ontology_column=None):
        params = {"type": "custom"}
        params["formula1"] = '=COUNTIF(${}:${},{}2)<2'.format(column, column, column)
        dv = DataValidation(**params)
        dv.error = 'Value must be unique'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Unique value".format(column_name)
        if self.unique_with:
            text += " Must be unique with column(s) {}".format(", ".join(self.unique_with))
        return text


def _validate_ontology(ontology, root_term=""):
    root_term_iri = ""
    if not ontology:
        return False
    base_path = "http://www.ebi.ac.uk/ols/api"
    sub_path = "/ontologies/{}".format(ontology.lower())
    r = requests.get(base_path + sub_path)
    if not r.status_code == 200:
        return False, root_term_iri
    if root_term:
        root_term_iri = _validate_ontological_term(root_term, ontology, return_uri=True)
    return True, root_term_iri


def _validate_ontological_term(term, ontology, root_term_iri="", return_uri=False):
    base_path = "http://www.ebi.ac.uk/ols/api/search"
    body = {
        "q": term,
        "ontology": ontology.lower(),
        "type": "class",
        "exact": True,
        "queryFields": ["label", "synonym"]
    }
    if root_term_iri:
        body["childrenOf"] = root_term_iri
    r = requests.get(base_path, params=body)
    res = r.json()
    if not res["response"]["numFound"] == 1:
        return False
    if return_uri:
        return res["response"]["docs"][0]["iri"]
    return True


def _get_ontological_terms(ontology, root_term_iri=""):
    size = 100
    terms = set()
    if root_term_iri:
        url = "http://www.ebi.ac.uk/ols/api/ontologies/{}/terms/{}/descendants?size={}".format(ontology, quote_plus(quote_plus(root_term_iri)), size)
    else:
        url = "http://www.ebi.ac.uk/ols/api/ontologies/{}/terms?size={}".format(ontology, size)

    r = requests.get(url)
    res = r.json()
    for term in res["_embedded"]["terms"]:
        terms.add(term["label"])
    while "next" in res["_links"]:
        url = res["_links"]["next"]["href"]
        r = requests.get(url)
        res = r.json()
        for term in res["_embedded"]["terms"]:
            terms.add(term["label"])

    return terms

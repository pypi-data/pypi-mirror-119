from pathlib import Path
from re import sub
import openpyxl
import pandas as pd


def read_title(file_path, skiprows=0, encoding='gbk', **kwargs):
    '''
    读取csv文件的表头
    '''
    title = pd.read_csv(file_path, skiprows=skiprows, nrows=0, header=kwargs.get('header', None), encoding=encoding)
    title = list(title.columns)
    return title


def get_files(path, filetype, include_subdir):
    '''
    根据输入的关键字搜索pth路径内的文件
    '''

    result = []
    path = Path(path)
    if not path.exists():
        raise ValueError('输入的路径不存在')
    if isinstance(filetype, str): filetype = [filetype]

    search_func = path.rglob if include_subdir else path.glob
    for key in filetype:
        result += [x for x in search_func('*' + key)]
    if not result:
        raise ValueError('未找到相关类型文件')
    return result


def read_excel(pth):
    '''
    读取excel里的每个sheet，yield每个sheet
    :param pth: 文件路径
    :return: 返回sheet
    '''
    wb = openpyxl.load_workbook(pth)
    for ws in wb.worksheets:
        if ws.sheet_state != 'hidden':
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                continue
            yield ws


def replace_invalid_char(s):
    '''
    在windows中不合法的文件路径内的字符替换成下划线
    :param title: 原字符
    :return: 替换后的字符
    '''
    rstr = r"[\/\\\:\*\?\"\<\>\|]."  # '/ \ : * ? " < > |'
    return sub(rstr, "_", s)  # 替换为下划线


def input_data(input_type='str', title='内容'):
    '''
    输入内容，并对返回值进行限定
    :param input_type: 对输入内容类型的限定
    :param title: 输入提示
    :return: 如果输入的类型符合限定则返回，否则继续输入
    '''
    while True:
        ipt = input(f'请输入{title}:')
        if ipt:
            try:
                return eval(input_type + f'("{ipt}")')
            except Exception as e:
                print(f'输入{title}错误，请重新输入')


def print_list_formating(print_string, step=4):
    '''
    将列表或字典格式化打印输出
    :param print_string: 列表或字典
    :param step: 打印列的数量
    :return: 无返回值
    '''
    print()
    if isinstance(print_string, list):
        print_string = [' ' + str(i + 1) + '、' + print_string[i] + ' ' for i in range(len(print_string))]
    elif isinstance(print_string, dict):
        print_string = [' ' + str(key) + ':' + str(value) + ' ' for key, value in print_string.items()]
    line_len = 0
    for i in range(0, step):
        max_len = 0
        for j in range(i, len(print_string), step):
            len_str = len(print_string[j].encode('gbk'))
            if max_len < len_str: max_len = len_str
        for j in range(i, len(print_string), step):
            len_str = len(print_string[j].encode('gbk'))
            print_string[j] = print_string[j] + ' ' * (max_len - len_str)
        line_len += max_len
    print('+' + '-' * line_len + '+')
    for i in range(0, len(print_string), step):
        list1 = print_string[i:i + step]
        s = ''.join(list1)
        slen = len(s.encode('gbk'))
        print('|' + s + ' ' * (line_len - slen) + '|')
    print('+' + '-' * line_len + '+')
    print()

from multiprocessing import Pool, cpu_count
from pathlib import Path
from .common import read_excel, replace_invalid_char, get_files
from csv import writer
import openpyxl
import csv


def _excel_trans_sheet_engine(save_path, workbook_path, skiprows=0):  # excel to sheet
    save_path = Path(save_path)
    if save_path.exists():
        dest_wb = openpyxl.load_workbook(save_path)
    else:
        dest_wb = openpyxl.Workbook()
    workbook_path = Path(workbook_path)
    name, stem = workbook_path.name, workbook_path.stem
    for sheet in read_excel(workbook_path):
        if sheet.max_row == 1 and sheet.max_row == 1 and sheet.cell(1, 1) is None:  # 空表不转换
            continue
        dest_ws_name = stem + '_' + sheet.title
        try:
            dest_ws = dest_wb.get_sheet_by_name(dest_ws_name)
        except:
            dest_ws = dest_wb.create_sheet(dest_ws_name)
        for row in sheet.iter_rows(min_row=skiprows, max_row=sheet.max_row, max_col=sheet.max_column):
            dest_ws.append([cell.value for cell in row])
    dest_wb.save(save_path)
    dest_wb.close()


def _excel_trans_excel_engine(save_path, workbook_path, encoding='gbk', skiprows=0):  # excel to csv
    workbook_path = Path(workbook_path)
    name, stem = workbook_path.name, workbook_path.stem
    file_list = []
    for sheet in read_excel(workbook_path):
        if sheet.max_row == 1 and sheet.max_row == 1 and sheet.cell(1, 1) is None:  # 空表不转换
            continue
        sheet_name = replace_invalid_char(sheet.title)
        _save_path = Path(save_path) / f'{stem}_{sheet_name}.csv'
        file_list.append(_save_path)
        print(f'转换文件：{_save_path}')
        with open(_save_path, 'w', encoding=encoding, newline='') as f:
            w = writer(f, dialect='excel')
            for row in sheet.rows:
                if skiprows > 0:
                    skiprows -= 1
                    continue
                w.writerow([cell.value for cell in row])
    return file_list


def _csv_trans_excel_engine(save_path, csv_path, encoding='gbk', skiprows=0):  # csv to excel
    csv_path = Path(csv_path)
    fstem, fname = csv_path.stem, csv_path.name
    fobj = open(csv_path, 'r', encoding=encoding)
    wb = openpyxl.load_workbook(save_path) if save_path.exists() else openpyxl.Workbook()
    try:
        ws = wb.get_sheet_by_name(fstem)
    except:
        ws = wb.create_sheet(fstem)
    lines = csv.reader(fobj)
    for _, line in enumerate(lines):
        if skiprows > 0:
            skiprows -= 1
            continue
        ws.append(line)
    wb.save(save_path)
    wb.close()
    fobj.close()


def excel_to_csv(save_path, workbook_path, encoding='gbk', skiprows=0):
    return _excel_trans_excel_engine(save_path, workbook_path, encoding=encoding, skiprows=skiprows)


def csv_to_excel(save_path, csv_path, encoding='gbk', skiprows=0):
    return _csv_trans_excel_engine(save_path, csv_path, encoding=encoding, skiprows=skiprows)


def excel_to_csv_in_path(save_path, file_path, types=None, encoding='gbk', include_subdir=False, skiprows=0):
    wb_paths = get_files(file_path, types, include_subdir)
    process_num = min(cpu_count(), len(wb_paths))
    result = []
    if wb_paths.__len__() >= process_num:
        p = Pool(processes=process_num)
        for wb_path in wb_paths:
            p.apply_async(excel_to_csv, args=(save_path, wb_path, encoding, skiprows), callback=result.extend)
        p.close()
        p.join()
    else:
        for wb_path in wb_paths:
            result.extend(excel_to_csv(save_path, wb_path, encoding, skiprows))
    return result


def csv_to_excel_in_path(save_path, file_path, types=None, encoding='gbk', include_subdir=False):
    csv_paths = get_files(file_path, types, include_subdir)
    process_num = min(cpu_count(), len(csv_paths))
    result = []
    if csv_paths.__len__() >= process_num:
        p = Pool(process_num)
        for csv_path in csv_paths:
            p.apply_async(csv_to_excel, args=(save_path, csv_path, encoding), callback=result.extend)
        p.close()
        p.join()
    else:
        for csv_path in csv_paths:
            result.extend(csv_to_excel(save_path, csv_path, encoding))
    return result
